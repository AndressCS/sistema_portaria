import tkinter as tk
from tkinter import messagebox, ttk, simpledialog, filedialog
import sqlite3
from datetime import datetime
import re
import hashlib
import secrets
import os
from PIL import Image, ImageTk
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import shutil
from tkinter import filedialog


DB_PATH = "portaria.db"

EMPRESAS = [
    "CARDEAL", "EBD", "ATAQ", "MARFIM", "DIA", "TORPEDO",
    "M&S", "TERRA BRASIL", "DRUGSTORE", "DISPAN",
    "MIX FARMA", "TEIXEIRA", "OUTROS"
]

# ===== Paleta Azul Institucional =====
COLORS = {
    "primary": "#0B3D91",
    "primary_dark": "#072B66",
    "accent": "#1E88E5",
    "bg": "#F4F6F9",
    "card": "#FFFFFF",
    "text": "#1F2937",
    "muted": "#6B7280",
    "success": "#1B5E20",
    "warning": "#F59E0B",
    "danger": "#B91C1C",
    "border": "#D6D9DE",
}


# ================= FUNÇÕES AUXILIARES =================

def forcar_maiusculo(var: tk.StringVar):
    texto = var.get()
    up = texto.upper()
    if texto != up:
        var.set(up)


def normalizar_placa(texto: str) -> str:
    return re.sub(r"[^A-Z0-9]", "", (texto or "").upper())


def placa_valida(placa: str) -> bool:
    p = normalizar_placa(placa)
    if re.fullmatch(r"[A-Z]{3}\d{4}", p):
        return True
    if re.fullmatch(r"[A-Z]{3}\d[A-Z]\d{2}", p):
        return True
    return False


def validar_placa_digitar(texto_digitado: str) -> bool:
    if texto_digitado is None:
        return True
    t = texto_digitado.upper()
    if not re.fullmatch(r"[A-Z0-9]*", t):
        return False
    if len(t) > 7:
        return False
    return True


def normalizar_telefone(texto: str) -> str:
    return re.sub(r"\D", "", texto or "")


def validar_telefone_digitar(texto_digitado: str) -> bool:
    if texto_digitado is None:
        return True
    if not re.fullmatch(r"\d*", texto_digitado):
        return False
    if len(texto_digitado) > 15:
        return False
    return True


def hash_senha(senha: str, salt_hex: str) -> str:
    dk = hashlib.pbkdf2_hmac(
        "sha256",
        senha.encode("utf-8"),
        bytes.fromhex(salt_hex),
        120_000
    )
    return dk.hex()


def usuario_valido(u: str) -> bool:
    return bool(re.fullmatch(r"[A-Z0-9_]{3,20}", (u or "").upper()))


def _asset_path(nome: str) -> str:
    base = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base, "assets", nome)


# ================= INTERFACE =================

def iniciar_tela():
    window = tk.Tk()
    window.title("Sistema de Portaria")
    window.state("zoomed")
    window.configure(bg=COLORS["bg"])

    # ===== ttk style =====
    style = ttk.Style(window)
    try:
        style.theme_use("clam")
    except Exception:
        pass

    style.configure("TFrame", background=COLORS["bg"])
    style.configure("Card.TFrame", background=COLORS["card"])
    style.configure("TLabel", background=COLORS["bg"], foreground=COLORS["text"])
    style.configure("Header.TFrame", background=COLORS["primary"])
    style.configure("Header.TLabel", background=COLORS["primary"], foreground="white")
    style.configure("CardTitle.TLabel", background=COLORS["card"], foreground=COLORS["primary"],
                    font=("Segoe UI", 11, "bold"))
    style.configure("SmallMuted.TLabel", background=COLORS["card"], foreground=COLORS["muted"])
    style.configure("TEntry", padding=6)
    style.configure("TCombobox", padding=4)

    # ===== DB =====
    con = sqlite3.connect(DB_PATH, timeout=10)
    con.execute("PRAGMA journal_mode=WAL;")
    con.execute("PRAGMA busy_timeout=10000;")
    cur = con.cursor()

    # ===== TABELAS =====
    cur.execute("""
    CREATE TABLE IF NOT EXISTS logs (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        data_hora TEXT NOT NULL,
        acao TEXT NOT NULL,
        entrada_id INTEGER,
        placa TEXT,
        destino TEXT,
        porteiro TEXT,
        detalhes TEXT
    )
    """)
    con.commit()

    cur.execute("""
    CREATE TABLE IF NOT EXISTS usuarios (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        usuario TEXT UNIQUE NOT NULL,
        nome TEXT NOT NULL,
        salt TEXT NOT NULL,
        senha_hash TEXT NOT NULL,
        role TEXT NOT NULL DEFAULT 'PORTEIRO'
    )
    """)
    con.commit()

    # ===== MIGRAÇÕES USUÁRIOS =====
    try:
        cur.execute("ALTER TABLE usuarios ADD COLUMN ativo INTEGER NOT NULL DEFAULT 1")
        con.commit()
    except Exception:
        pass

    # ===== MIGRAÇÕES ENTRADAS =====
    def garantir_coluna(sql):
        try:
            cur.execute(sql)
            con.commit()
        except Exception:
            pass

    cur.execute("""
    CREATE TABLE IF NOT EXISTS entradas (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        motorista TEXT,
        placa TEXT,
        telefone TEXT,
        fornecedor TEXT,
        destino TEXT,
        data_hora TEXT,
        porteiro TEXT,
        saida TEXT
    )
    """)
    con.commit()

    garantir_coluna("ALTER TABLE entradas ADD COLUMN telefone TEXT")
    garantir_coluna("ALTER TABLE entradas ADD COLUMN fornecedor TEXT")
    garantir_coluna("ALTER TABLE entradas ADD COLUMN destino TEXT")
    garantir_coluna("ALTER TABLE entradas ADD COLUMN data_hora TEXT")
    garantir_coluna("ALTER TABLE entradas ADD COLUMN porteiro TEXT")
    garantir_coluna("ALTER TABLE entradas ADD COLUMN saida TEXT")

    # ===== sessão =====
    usuario_logado = {"usuario": None, "nome": None, "role": None}
    baixas_em_andamento = set()
    desfazer_em_andamento = set()

    # ===== status bar =====
    status_var = tk.StringVar(value="Pronto.")

    def set_status(msg: str, kind: str = "info"):
        if kind == "success":
            bg = COLORS["success"]
        elif kind == "warning":
            bg = COLORS["warning"]
        elif kind == "error":
            bg = COLORS["danger"]
        else:
            bg = COLORS["primary_dark"]

        status_label.configure(bg=bg)
        status_var.set(msg)

    def on_close():
        try:
            con.close()
        except Exception:
            pass
        window.destroy()

    window.protocol("WM_DELETE_WINDOW", on_close)

    def registrar_log(acao, entrada_id=None, placa=None, destino=None, porteiro=None, detalhes=""):
        porteiro = (porteiro or "").strip() or "NÃO INFORMADO"
        cur.execute("""
            INSERT INTO logs (data_hora, acao, entrada_id, placa, destino, porteiro, detalhes)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """, (
            datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
            acao, entrada_id, placa, destino, porteiro, detalhes
        ))
        con.commit()

    def criar_usuario_se_nao_existir(usuario: str, nome: str, senha: str, role: str = "PORTEIRO"):
        usuario = (usuario or "").strip().upper()
        nome = (nome or "").strip().upper()
        role = (role or "").strip().upper()

        cur.execute("SELECT 1 FROM usuarios WHERE usuario = ?", (usuario,))
        if cur.fetchone():
            return

        salt = secrets.token_hex(16)
        senha_hash = hash_senha(senha, salt)
        cur.execute("""
            INSERT INTO usuarios (usuario, nome, salt, senha_hash, role, ativo)
            VALUES (?, ?, ?, ?, ?, 1)
        """, (usuario, nome, salt, senha_hash, role))
        con.commit()

    criar_usuario_se_nao_existir("ADMIN", "ADMIN", "ADMIN123", "ADMIN")

    # ================= MODAL PREMIUM =================

    def confirmar_modal(
        titulo: str,
        mensagem: str,
        texto_ok="CONFIRMAR",
        texto_cancel="CANCELAR",
        tipo="warning",
        detalhes_linhas=None
    ):
        """
        Modal premium de confirmação com:
        - overlay escurecido
        - animação fade-in
        - Enter confirma / Esc cancela
        - bloco de detalhes (placa/motorista/destino etc.)
        Retorna True ou False.
        """
        detalhes_linhas = detalhes_linhas or []

        # Overlay (escurece o fundo)
        overlay = tk.Toplevel(window)
        overlay.overrideredirect(True)
        overlay.configure(bg="black")
        overlay.attributes("-alpha", 0.0)  # animado
        overlay.lift()
        overlay.grab_set()

        window.update_idletasks()
        x = window.winfo_rootx()
        y = window.winfo_rooty()
        w = window.winfo_width()
        h = window.winfo_height()
        overlay.geometry(f"{w}x{h}+{x}+{y}")

        # Modal
        modal = tk.Toplevel(window)
        modal.title(titulo)
        modal.configure(bg=COLORS["bg"])
        modal.resizable(False, False)
        modal.transient(window)
        modal.lift()
        modal.attributes("-alpha", 0.0)  # animado
        modal.grab_set()

        # Ícone + cor
        if tipo == "danger":
            cor = COLORS["danger"]
            icone = "⛔"
        elif tipo == "info":
            cor = COLORS["accent"]
            icone = "ℹ️"
        else:
            cor = COLORS["warning"]
            icone = "⚠️"

        # Card
        card = ttk.Frame(modal, style="Card.TFrame", padding=18)
        card.pack(fill="both", expand=True)
        card.configure(relief="solid")
        card["borderwidth"] = 1

        # Faixa superior
        top = ttk.Frame(card, style="Card.TFrame")
        top.pack(fill="x")

        pill = tk.Label(
            top,
            text=icone,
            bg=COLORS["card"],
            fg=cor,
            font=("Segoe UI", 22, "bold")
        )
        pill.pack(side="left", padx=(0, 10))

        title_lbl = tk.Label(
            top,
            text=titulo,
            bg=COLORS["card"],
            fg=COLORS["primary"],
            font=("Segoe UI", 14, "bold")
        )
        title_lbl.pack(side="left", pady=(2, 0))

        ttk.Separator(card).pack(fill="x", pady=12)

        msg_lbl = tk.Label(
            card,
            text=mensagem,
            bg=COLORS["card"],
            fg=COLORS["text"],
            font=("Segoe UI", 11),
            justify="left",
            wraplength=560
        )
        msg_lbl.pack(anchor="w")

        # Detalhes (caixa)
        if detalhes_linhas:
            box = ttk.Frame(card, style="Card.TFrame", padding=10)
            box.pack(side="left", padx=(0, 24), pady=10)
            box.configure(relief="solid")
            box["borderwidth"] = 1

            tk.Label(
                box,
                text="DETALHES",
                bg=COLORS["card"],
                fg=COLORS["muted"],
                font=("Segoe UI", 9, "bold")
            ).pack(anchor="w")

            for ln in detalhes_linhas:
                tk.Label(
                    box,
                    text=ln,
                    bg=COLORS["card"],
                    fg=COLORS["text"],
                    font=("Segoe UI", 10),
                    justify="left",
                    anchor="w"
                ).pack(anchor="w", pady=1)

        # Botões
        btns = ttk.Frame(card, style="Card.TFrame")
        btns.pack(fill="x", pady=(16, 0))

        resultado = {"ok": False}

        def fechar(ok: bool):
            resultado["ok"] = ok
            try:
                modal.grab_release()
            except Exception:
                pass
            try:
                overlay.grab_release()
            except Exception:
                pass
            overlay.destroy()
            modal.destroy()

        def on_escape(event=None):
            fechar(False)

        def on_enter(event=None):
            fechar(True)

        modal.bind("<Escape>", on_escape)

        # Botão cancelar
        btn_cancel = tk.Button(
            btns,
            text=texto_cancel,
            command=lambda: fechar(False),
            bd=0,
            padx=14,
            pady=10,
            bg="#374151",
            fg="white",
            activebackground=COLORS["primary_dark"],
            activeforeground="white",
            cursor="hand2",
            font=("Segoe UI", 10, "bold")
        )
        btn_cancel.pack(side="right", padx=(8, 0))

        # Botão confirmar
        btn_ok = tk.Button(
            btns,
            text=texto_ok,
            command=lambda: fechar(True),
            bd=0,
            padx=14,
            pady=10,
            bg=cor,
            fg="white",
            activebackground=COLORS["primary_dark"],
            activeforeground="white",
            cursor="hand2",
            font=("Segoe UI", 10, "bold")
        )
        btn_ok.pack(side="right")

        # Enter confirma (global no modal)
        modal.bind("<Return>", on_enter)

        # Centralizar modal (tamanho real do conteúdo)
        modal.update_idletasks()

        # largura fixa “premium” (mantém visual consistente)
        mw = 680
        modal.geometry(f"{mw}x1")  # altura mínima provisória para o Tk recalcular

        modal.update_idletasks()
        req_w = max(mw, modal.winfo_reqwidth())
        req_h = modal.winfo_reqheight()

        # limita ao tamanho da tela (evita cortar em telas pequenas / escala alta)
        screen_w = modal.winfo_screenwidth()
        screen_h = modal.winfo_screenheight()

        max_w = int(screen_w * 0.85)
        max_h = int(screen_h * 0.85)

        final_w = min(req_w, max_w)
        final_h = min(req_h, max_h)

        sx = window.winfo_rootx()
        sy = window.winfo_rooty()
        sw = window.winfo_width()
        sh = window.winfo_height()

        mx = sx + (sw // 2) - (final_w // 2)
        my = sy + (sh // 2) - (final_h // 2)

        # garante que não “escape” da tela
        mx = max(10, min(mx, screen_w - final_w - 10))
        my = max(10, min(my, screen_h - final_h - 10))

        modal.geometry(f"{final_w}x{final_h}+{mx}+{my}")

        # Animação fade-in
        def fade(step=0):
            # overlay até 0.22, modal até 1.0
            o = min(0.22, step * 0.02)
            m = min(1.0, step * 0.08)
            try:
                overlay.attributes("-alpha", o)
                modal.attributes("-alpha", m)
            except Exception:
                return
            if m < 1.0:
                modal.after(12, lambda: fade(step + 1))

        fade()

        # foco
        btn_ok.focus_set()
        modal.wait_window()
        return resultado["ok"]

    # ================== LAYOUT PRINCIPAL ==================
    root_container = ttk.Frame(window)
    root_container.pack(fill="both", expand=True)

    # ===== Header =====
    header = ttk.Frame(root_container, style="Header.TFrame", padding=(14, 10))
    header.pack(fill="x")

    ttk.Label(header, text="SISTEMA DE PORTARIA", style="Header.TLabel",
              font=("Segoe UI", 14, "bold")).pack(side="left")

    user_label_var = tk.StringVar(value="Usuário: (não logado)")
    ttk.Label(header, textvariable=user_label_var, style="Header.TLabel",
              font=("Segoe UI", 10)).pack(side="left", padx=(12, 0))

    header_btns = ttk.Frame(header, style="Header.TFrame")
    header_btns.pack(side="right")

    def header_button(parent, text, cmd, color=None):
        return tk.Button(
            parent,
            text=text,
            command=cmd,
            bd=0,
            padx=12,
            pady=6,
            fg="white",
            bg=color or COLORS["accent"],
            activebackground=COLORS["primary_dark"],
            activeforeground="white",
            cursor="hand2",
            font=("Segoe UI", 9, "bold")
        )

    # ===== Body =====
    body = ttk.Frame(root_container, padding=(14, 12))
    body.pack(fill="both", expand=True)

    left = ttk.Frame(body)
    left.pack(side="left", fill="y", padx=(0, 10))

    right = ttk.Frame(body)
    right.pack(side="left", fill="both", expand=True)

    # ===== Cards (Form / Busca) =====
    form_card = ttk.Frame(left, style="Card.TFrame", padding=12)
    form_card.pack(fill="x", pady=(0, 10))
    form_card.configure(relief="solid")
    form_card["borderwidth"] = 1

    ttk.Label(form_card, text="Cadastro de Entrada", style="CardTitle.TLabel").pack(anchor="w", pady=(0, 10))

    var_motorista = tk.StringVar()
    var_placa = tk.StringVar()
    var_telefone = tk.StringVar()
    var_fornecedor = tk.StringVar()
    var_destino = tk.StringVar()
    var_porteiro = tk.StringVar()

    var_motorista.trace_add("write", lambda *a: forcar_maiusculo(var_motorista))
    var_fornecedor.trace_add("write", lambda *a: forcar_maiusculo(var_fornecedor))
    var_porteiro.trace_add("write", lambda *a: forcar_maiusculo(var_porteiro))

    form_grid = ttk.Frame(form_card, style="Card.TFrame")
    form_grid.pack(fill="x")
    form_grid.grid_columnconfigure(1, weight=1)

    def mk_label(txt, r):
        ttk.Label(form_grid, text=txt, background=COLORS["card"], foreground=COLORS["text"],
                  font=("Segoe UI", 9, "bold")).grid(row=r, column=0, sticky="w", pady=4)

    def mk_entry(var, r, validate=None):
        e = ttk.Entry(form_grid, textvariable=var, width=32)
        if validate:
            e.configure(validate="key", validatecommand=validate)
        e.grid(row=r, column=1, sticky="we", pady=4)
        return e

    mk_label("Motorista", 0)
    entry_motorista = mk_entry(var_motorista, 0)

    mk_label("Placa", 1)
    vcmd_placa = (window.register(validar_placa_digitar), "%P")
    entry_placa = mk_entry(var_placa, 1, validate=vcmd_placa)

    mk_label("Telefone", 2)
    vcmd_tel = (window.register(validar_telefone_digitar), "%P")
    entry_tel = mk_entry(var_telefone, 2, validate=vcmd_tel)

    mk_label("Fornecedor", 3)
    entry_forn = mk_entry(var_fornecedor, 3)

    mk_label("Destino", 4)
    combo_destino = ttk.Combobox(form_grid, textvariable=var_destino, values=EMPRESAS, state="readonly", width=30)
    combo_destino.grid(row=4, column=1, sticky="we", pady=4)

    mk_label("Porteiro", 5)
    entry_porteiro = mk_entry(var_porteiro, 5)

    # ===== Busca =====
    busca_card = ttk.Frame(left, style="Card.TFrame", padding=12)
    busca_card.pack(fill="x", pady=(0, 10))
    busca_card.configure(relief="solid")
    busca_card["borderwidth"] = 1

    ttk.Label(busca_card, text="Busca Rápida", style="CardTitle.TLabel").pack(anchor="w", pady=(0, 10))
    ttk.Label(busca_card, text="Pesquisar por placa ou empresa:", style="SmallMuted.TLabel").pack(anchor="w")

    var_busca = tk.StringVar()
    var_busca.trace_add("write", lambda *a: (forcar_maiusculo(var_busca), carregar_blocos()))
    entry_busca = ttk.Entry(busca_card, textvariable=var_busca, width=30)
    entry_busca.pack(fill="x", pady=(6, 0))

    # ===== Botões principais =====
    btns_card = ttk.Frame(left)
    btns_card.pack(fill="x", pady=(0, 10))

    def main_button(parent, text, cmd, color):
        return tk.Button(
            parent,
            text=text,
            command=cmd,
            bd=0,
            padx=12,
            pady=10,
            fg="white",
            bg=color,
            activebackground=COLORS["primary_dark"],
            activeforeground="white",
            cursor="hand2",
            font=("Segoe UI", 10, "bold")
        )

    # ================= LOGIN / USUÁRIOS =================

    def abrir_login():
        login = tk.Toplevel(window)
        login.title("Login - Sistema de Portaria")
        login.state("zoomed")
        login.configure(bg=COLORS["bg"])
        login.grab_set()

        def bloquear_sem_login():
            if usuario_logado.get("usuario") is None:
                try:
                    window.destroy()
                except Exception:
                    pass

        login.protocol("WM_DELETE_WINDOW", bloquear_sem_login)

        # ===== Helpers de imagem (sem Canvas, sem watermark) =====
        def carregar_imagem(nome_arquivo, tamanho=None):
            caminho = _asset_path(nome_arquivo)
            try:
                img = Image.open(caminho).convert("RGBA")
                if tamanho:
                    img = img.resize(tamanho, Image.LANCZOS)
                return ImageTk.PhotoImage(img)
            except Exception as e:
                print("Erro ao carregar imagem:", e)
                print("Tentou abrir em:", caminho)
                return None

        # ===== Wrapper geral =====
        wrapper = ttk.Frame(login, padding=28)
        wrapper.pack(fill="both", expand=True)

        wrapper.grid_columnconfigure(0, weight=6)  # hero
        wrapper.grid_columnconfigure(1, weight=4)  # card
        wrapper.grid_rowconfigure(0, weight=1)

        # ===== Painel HERO (esquerda) =====
        left_panel = ttk.Frame(wrapper, style="Card.TFrame", padding=22)
        left_panel.grid(row=0, column=0, sticky="nsew", padx=(0, 16))
        left_panel.configure(relief="flat")
        left_panel["borderwidth"] = 0

        # Topo do hero
        tk.Label(
            left_panel,
            text="SISTEMA DE PORTARIA",
            bg=COLORS["card"],
            fg=COLORS["primary"],
            font=("Segoe UI", 28, "bold")
        ).pack(anchor="w", pady=(2, 2))

        tk.Label(
            left_panel,
            text="Controle de entradas, baixas, histórico, logs e relatórios — de forma rápida e segura.",
            bg=COLORS["card"],
            fg=COLORS["muted"],
            font=("Segoe UI", 12)
        ).pack(anchor="w", pady=(0, 16))

        # Header "DAS EMPRESAS" (compacto e alinhado com as logos)
        header_emp = ttk.Frame(left_panel, style="Card.TFrame")
        header_emp.pack(fill="x", pady=(2, 8))

        tk.Label(
            header_emp,
            text="Acesso dedicado para operações e registros de portaria.",
            bg=COLORS["card"],
            fg="#6B7280",
            font=("Segoe UI", 10)
        ).pack(anchor="w", pady=(4, 0))

        ttk.Separator(left_panel).pack(fill="x", pady=(6, 14))

        # Logos grandes (sem transparência, sem watermark)
        logos_row = ttk.Frame(left_panel, style="Card.TFrame")
        logos_row.pack(fill="x")

        login._logo_verzani = carregar_imagem("verzani.png", (420, 170))
        login._logo_ceipe = carregar_imagem("ceipe.png", (420, 170))

        def logo_box(parent, img, fallback_text):
            box = ttk.Frame(parent, style="Card.TFrame", padding=12)
            box.pack(side="left", padx=(0, 14))
            box.configure(relief="flat")
            box["borderwidth"] = 0

            if img:
                tk.Label(box, image=img, bg=COLORS["card"]).pack()
            else:
                tk.Label(box, text=fallback_text, bg=COLORS["card"], fg=COLORS["muted"],
                         font=("Segoe UI", 16, "bold")).pack()

        logo_box(logos_row, login._logo_verzani, "VERZANI")
        logo_box(logos_row, login._logo_ceipe, "CEIPE")

        ttk.Separator(left_panel).pack(fill="x", pady=16)

        # Empresas em GRID fixo (não empurra / não treme)
        tk.Label(
            left_panel,
            text="EMPRESAS / DESTINOS ATENDIDOS",
            bg=COLORS["card"],
            fg=COLORS["primary"],
            font=("Segoe UI", 12, "bold")
        ).pack(anchor="w")

        chips_wrap = ttk.Frame(left_panel, style="Card.TFrame")
        chips_wrap.pack(fill="x", pady=(10, 0))

        cols = 4  # estável e bonito
        for i in range(cols):
            chips_wrap.grid_columnconfigure(i, weight=1)

        r = 0
        c = 0
        for nome in EMPRESAS:
            chip = tk.Label(
                chips_wrap,
                text=nome,
                bg="#F3F4F6",  # cinza claro
                fg="#1F2937",  # texto escuro
                padx=10,
                pady=6,
                font=("Segoe UI", 8, "bold"),
                bd=0,
                relief="solid"
            )
            chip.grid(row=r, column=c, sticky="w", padx=6, pady=6)

            c += 1
            if c >= cols:
                c = 0
                r += 1

        # ===== Painel LOGIN (direita) =====
        right_panel = ttk.Frame(wrapper, style="Card.TFrame", padding=0)
        right_panel.grid(row=0, column=1, sticky="nsew")
        right_panel.configure(relief="solid")
        right_panel["borderwidth"] = 1

        # Borda dupla para “sombra fake” (premium e leve)
        inner_shadow = tk.Frame(right_panel, bg="#E5E7EB")  # cinza claro
        inner_shadow.pack(fill="both", expand=True, padx=12, pady=12)

        inner = tk.Frame(inner_shadow, bg=COLORS["card"])
        inner.pack(fill="both", expand=True, padx=2, pady=2)

        # Conteúdo do card
        tk.Label(
            inner,
            text="Acesso ao sistema",
            bg=COLORS["card"],
            fg=COLORS["text"],
            font=("Segoe UI", 18, "bold")
        ).pack(anchor="w", padx=18, pady=(18, 6))

        tk.Label(
            inner,
            text="Entre com seu usuário e senha.",
            bg=COLORS["card"],
            fg=COLORS["muted"],
            font=("Segoe UI", 11)
        ).pack(anchor="w", padx=18, pady=(0, 16))

        # Form
        form = tk.Frame(inner, bg=COLORS["card"])
        form.pack(fill="x", padx=18)

        var_user = tk.StringVar()
        var_user.trace_add("write", lambda *a: forcar_maiusculo(var_user))
        var_pass = tk.StringVar()
        var_show = tk.BooleanVar(value=False)

        def label(txt):
            tk.Label(form, text=txt, bg=COLORS["card"], fg=COLORS["muted"],
                     font=("Segoe UI", 10, "bold")).pack(anchor="w", pady=(10, 6))

        label("Usuário")
        e_user = ttk.Entry(form, textvariable=var_user)
        e_user.pack(fill="x", ipady=7)

        label("Senha")
        pass_row = tk.Frame(form, bg=COLORS["card"])
        pass_row.pack(fill="x")

        e_pass = ttk.Entry(pass_row, textvariable=var_pass, show="*")
        e_pass.pack(side="left", fill="x", expand=True, ipady=7)

        def toggle_senha():
            e_pass.configure(show="" if var_show.get() else "*")

        ttk.Checkbutton(pass_row, text="Mostrar", variable=var_show, command=toggle_senha) \
            .pack(side="left", padx=(10, 0))

        lbl_status = tk.Label(inner, text="", bg=COLORS["card"], fg=COLORS["danger"],
                              font=("Segoe UI", 10, "bold"))
        lbl_status.pack(anchor="w", padx=18, pady=(12, 0))

        def tentar_login():
            u = (var_user.get() or "").strip().upper()
            p = (var_pass.get() or "").strip()

            if not u or not p:
                lbl_status.config(text="Informe usuário e senha.")
                return

            cur.execute("""
                        SELECT usuario, nome, salt, senha_hash, role, ativo
                        FROM usuarios
                        WHERE usuario = ?
                        """, (u,))
            row = cur.fetchone()
            if not row:
                lbl_status.config(text="Usuário não encontrado.")
                return

            usuario_db, nome_db, salt_db, hash_db, role_db, ativo_db = row
            if int(ativo_db) != 1:
                lbl_status.config(text="Usuário desativado.")
                return

            if hash_senha(p, salt_db) != hash_db:
                lbl_status.config(text="Senha incorreta.")
                return

            usuario_logado["usuario"] = usuario_db
            usuario_logado["nome"] = nome_db
            usuario_logado["role"] = role_db

            var_porteiro.set(nome_db)
            try:
                entry_porteiro.configure(state="disabled")
            except Exception:
                pass

            user_label_var.set(f"Usuário: {nome_db} ({role_db})")
            set_status("Login realizado.", "success")
            login.destroy()

        # Botão principal
        btn_login = tk.Button(
            inner,
            text="ENTRAR",
            command=tentar_login,
            bd=0,
            padx=16,
            pady=14,
            bg=COLORS["primary"],
            fg="white",
            activebackground=COLORS["primary_dark"],
            activeforeground="white",
            cursor="hand2",
            font=("Segoe UI", 12, "bold")
        )
        btn_login.pack(fill="x", padx=18, pady=(18, 0))

        # Rodapé com dicas
        tk.Label(
            inner,
            text="Enter para entrar • Esc para limpar senha",
            bg=COLORS["card"],
            fg=COLORS["muted"],
            font=("Segoe UI", 9)
        ).pack(anchor="w", padx=18, pady=(12, 18))

        # Teclas
        def ao_enter(event=None):
            tentar_login()

        def ao_esc(event=None):
            var_pass.set("")
            lbl_status.config(text="")

        e_user.bind("<Return>", ao_enter)
        e_pass.bind("<Return>", ao_enter)
        login.bind("<Escape>", ao_esc)

        e_user.focus_set()

    def logout():
        nome_atual = usuario_logado.get("nome") or ""
        usuario_atual = usuario_logado.get("usuario") or ""

        if not confirmar_modal(
            "Trocar usuário",
            "Deseja realmente trocar o usuário logado?\n\nVocê retornará para a tela de login.",
            texto_ok="TROCAR",
            texto_cancel="CANCELAR",
            tipo="info",
            detalhes_linhas=[
                f"USUÁRIO: {usuario_atual}",
                f"NOME: {nome_atual}"
            ] if nome_atual or usuario_atual else []
        ):
            return

        if usuario_logado.get("nome"):
            registrar_log("LOGOUT", porteiro=usuario_logado.get("nome"),
                          detalhes=f"TROCOU USUÁRIO (de {usuario_logado.get('usuario')})")

        usuario_logado["usuario"] = None
        usuario_logado["nome"] = None
        usuario_logado["role"] = None

        entry_porteiro.configure(state="normal")
        var_porteiro.set("")
        user_label_var.set("Usuário: (não logado)")

        abrir_login()

    def abrir_cadastro_usuarios():
        if usuario_logado.get("role") != "ADMIN":
            messagebox.showwarning("Acesso negado", "Apenas ADMIN pode gerenciar usuários.")
            return

        win = tk.Toplevel(window)
        win.title("Cadastro de Usuários")
        win.state("zoomed")
        win.configure(bg=COLORS["bg"])

        container = ttk.Frame(win, padding=14)
        container.pack(fill="both", expand=True)

        card = ttk.Frame(container, style="Card.TFrame", padding=12)
        card.pack(fill="both", expand=True)
        card.configure(relief="solid")
        card["borderwidth"] = 1

        ttk.Label(card, text="Usuários do Sistema", style="CardTitle.TLabel").pack(anchor="w", pady=(0, 10))

        frame_lista = ttk.Frame(card, style="Card.TFrame")
        frame_lista.pack(fill="both", expand=True)

        cols = ("usuario", "nome", "role", "ativo")
        tree = ttk.Treeview(frame_lista, columns=cols, show="headings")
        tree.heading("usuario", text="USUÁRIO")
        tree.heading("nome", text="NOME")
        tree.heading("role", text="PERFIL")
        tree.heading("ativo", text="ATIVO")
        tree.column("usuario", width=140, anchor="w")
        tree.column("nome", width=260, anchor="w")
        tree.column("role", width=120, anchor="center")
        tree.column("ativo", width=80, anchor="center")

        frame_lista.grid_rowconfigure(0, weight=1)
        frame_lista.grid_columnconfigure(0, weight=1)
        tree.grid(row=0, column=0, sticky="nsew")

        scroll_y = ttk.Scrollbar(frame_lista, orient="vertical", command=tree.yview)
        scroll_y.grid(row=0, column=1, sticky="ns")
        scroll_x = ttk.Scrollbar(frame_lista, orient="horizontal", command=tree.xview)
        scroll_x.grid(row=1, column=0, sticky="ew")
        tree.configure(yscrollcommand=scroll_y.set, xscrollcommand=scroll_x.set)

        btns = ttk.Frame(card, style="Card.TFrame")
        btns.pack(fill="x", pady=(10, 0))

        def carregar_usuarios():
            for item in tree.get_children():
                tree.delete(item)
            cur.execute("SELECT usuario, nome, role, ativo FROM usuarios ORDER BY usuario")
            for u, n, r, a in cur.fetchall():
                tree.insert("", "end", values=(u, n, r, "SIM" if int(a) == 1 else "NÃO"))

        def get_sel():
            sel = tree.selection()
            if not sel:
                return None
            return tree.item(sel[0], "values")

        def novo_usuario():
            u = simpledialog.askstring("Novo usuário", "Usuário (A-Z, 0-9, _), 3-20:", parent=win)
            if u is None:
                return
            u = u.strip().upper()
            if not usuario_valido(u):
                messagebox.showwarning("Usuário inválido", "Use apenas A-Z, 0-9 e _. Tamanho 3 a 20.")
                return

            n = simpledialog.askstring("Novo usuário", "Nome completo:", parent=win)
            if n is None:
                return
            n = n.strip().upper()
            if not n:
                messagebox.showwarning("Nome inválido", "Informe o nome.")
                return

            senha = simpledialog.askstring("Novo usuário", "Senha inicial:", show="*", parent=win)
            if senha is None:
                return
            if len(senha) < 4:
                messagebox.showwarning("Senha fraca", "A senha deve ter pelo menos 4 caracteres.")
                return

            role = simpledialog.askstring("Novo usuário", "Perfil (PORTEIRO ou ADMIN):", parent=win)
            if role is None:
                return
            role = role.strip().upper()
            if role not in ("PORTEIRO", "ADMIN"):
                messagebox.showwarning("Perfil inválido", "Use PORTEIRO ou ADMIN.")
                return

            try:
                salt = secrets.token_hex(16)
                senha_hash = hash_senha(senha, salt)
                cur.execute("""
                    INSERT INTO usuarios (usuario, nome, salt, senha_hash, role, ativo)
                    VALUES (?, ?, ?, ?, ?, 1)
                """, (u, n, salt, senha_hash, role))
                con.commit()

                registrar_log("USER_CREATE", porteiro=usuario_logado.get("nome"),
                              detalhes=f"CRIADO: usuario={u} nome={n} role={role}")
                set_status(f"Usuário {u} criado.", "success")
                carregar_usuarios()
            except sqlite3.IntegrityError:
                messagebox.showwarning("Já existe", "Este usuário já existe.")
            except Exception as e:
                messagebox.showerror("Erro", str(e))

        def reset_senha():
            sel = get_sel()
            if not sel:
                messagebox.showwarning("Selecione", "Selecione um usuário.")
                return
            u, n, r, a = sel

            senha = simpledialog.askstring("Resetar senha", f"Nova senha para {u}:", show="*", parent=win)
            if senha is None:
                return
            if len(senha) < 4:
                messagebox.showwarning("Senha fraca", "A senha deve ter pelo menos 4 caracteres.")
                return

            salt = secrets.token_hex(16)
            senha_hash = hash_senha(senha, salt)
            cur.execute("UPDATE usuarios SET salt=?, senha_hash=? WHERE usuario=?", (salt, senha_hash, u))
            con.commit()

            registrar_log("USER_RESET_PASS", porteiro=usuario_logado.get("nome"), detalhes=f"RESET SENHA: usuario={u}")
            set_status(f"Senha do usuário {u} atualizada.", "success")

        def trocar_role():
            sel = get_sel()
            if not sel:
                messagebox.showwarning("Selecione", "Selecione um usuário.")
                return
            u, n, r, a = sel

            novo = simpledialog.askstring("Trocar perfil", f"Novo perfil para {u} (PORTEIRO/ADMIN):", parent=win)
            if novo is None:
                return
            novo = novo.strip().upper()
            if novo not in ("PORTEIRO", "ADMIN"):
                messagebox.showwarning("Perfil inválido", "Use PORTEIRO ou ADMIN.")
                return

            cur.execute("UPDATE usuarios SET role=? WHERE usuario=?", (novo, u))
            con.commit()

            registrar_log("USER_ROLE", porteiro=usuario_logado.get("nome"),
                          detalhes=f"ROLE: usuario={u} de={r} para={novo}")
            set_status(f"Perfil de {u} atualizado.", "success")
            carregar_usuarios()

        def ativar_desativar():
            sel = get_sel()
            if not sel:
                messagebox.showwarning("Selecione", "Selecione um usuário.")
                return
            u, n, r, a = sel

            if u == (usuario_logado.get("usuario") or ""):
                messagebox.showwarning("Não permitido", "Você não pode desativar o usuário logado.")
                return

            ativo_atual = 1 if a == "SIM" else 0
            novo_ativo = 0 if ativo_atual == 1 else 1
            acao_txt = "DESATIVAR" if novo_ativo == 0 else "ATIVAR"

            if not confirmar_modal(
                "Confirmar alteração",
                f"Deseja {acao_txt} o usuário selecionado?",
                texto_ok=acao_txt,
                texto_cancel="CANCELAR",
                tipo="warning",
                detalhes_linhas=[f"USUÁRIO: {u}", f"NOME: {n}", f"PERFIL: {r}"]
            ):
                return

            cur.execute("UPDATE usuarios SET ativo=? WHERE usuario=?", (novo_ativo, u))
            con.commit()

            registrar_log("USER_ACTIVE", porteiro=usuario_logado.get("nome"),
                          detalhes=f"ATIVO: usuario={u} para={'SIM' if novo_ativo == 1 else 'NÃO'}")
            set_status(f"Usuário {u} atualizado.", "success")
            carregar_usuarios()

        tk.Button(btns, text="NOVO", command=novo_usuario, bd=0, padx=10, pady=8,
                  bg=COLORS["primary"], fg="white", font=("Segoe UI", 9, "bold")).pack(side="left", padx=6)
        tk.Button(btns, text="RESET SENHA", command=reset_senha, bd=0, padx=10, pady=8,
                  bg=COLORS["accent"], fg="white", font=("Segoe UI", 9, "bold")).pack(side="left", padx=6)
        tk.Button(btns, text="TROCAR PERFIL", command=trocar_role, bd=0, padx=10, pady=8,
                  bg=COLORS["primary_dark"], fg="white", font=("Segoe UI", 9, "bold")).pack(side="left", padx=6)
        tk.Button(btns, text="ATIVAR/DESATIVAR", command=ativar_desativar, bd=0, padx=10, pady=8,
                  bg=COLORS["warning"], fg="white", font=("Segoe UI", 9, "bold")).pack(side="left", padx=6)
        tk.Button(btns, text="ATUALIZAR", command=carregar_usuarios, bd=0, padx=10, pady=8,
                  bg="#374151", fg="white", font=("Segoe UI", 9, "bold")).pack(side="right", padx=6)

        carregar_usuarios()

    # ================= AUTO-PREENCHIMENTO POR PLACA =================
    def autopreencher_por_placa():
        placa_norm = normalizar_placa(var_placa.get())
        if var_placa.get() != placa_norm:
            var_placa.set(placa_norm)
            return

        if len(placa_norm) < 7:
            return

        try:
            cur.execute("""
                SELECT motorista, telefone, fornecedor
                FROM entradas
                WHERE placa = ?
                ORDER BY id DESC
                LIMIT 1
            """, (placa_norm,))
            row = cur.fetchone()
            if not row:
                return

            motorista_db, telefone_db, fornecedor_db = row

            if not var_motorista.get():
                var_motorista.set(motorista_db or "")
            if not var_telefone.get():
                var_telefone.set(telefone_db or "")
            if not var_fornecedor.get():
                var_fornecedor.set(fornecedor_db or "")
        except Exception:
            pass

    var_placa.trace_add("write", lambda *a: autopreencher_por_placa())

    # ================= FUNÇÕES DE NEGÓCIO =================
    def limpar_campos():
        var_motorista.set("")
        var_placa.set("")
        var_telefone.set("")
        var_fornecedor.set("")
        var_destino.set("")
        set_status("Campos limpos.", "info")

    def registrar_entrada():
        try:
            placa_norm = normalizar_placa(var_placa.get())
            var_placa.set(placa_norm)

            tel_norm = normalizar_telefone(var_telefone.get())
            var_telefone.set(tel_norm)

            if not placa_valida(placa_norm):
                messagebox.showwarning("Placa inválida", "Formato: ABC1234 ou ABC1D23")
                set_status("Placa inválida.", "warning")
                return

            if not all([
                var_motorista.get().strip(),
                var_placa.get().strip(),
                var_telefone.get().strip(),
                var_fornecedor.get().strip(),
                var_destino.get().strip(),
                var_porteiro.get().strip()
            ]):
                messagebox.showwarning("Atenção", "Preencha todos os campos.")
                set_status("Preencha todos os campos.", "warning")
                return

            cur.execute("SELECT id FROM entradas WHERE placa = ? AND saida IS NULL", (var_placa.get(),))
            if cur.fetchone():
                messagebox.showwarning("Placa já registrada", "Este veículo já possui ENTRADA ATIVA.")
                set_status("Placa já possui entrada ativa.", "warning")
                return

            cur.execute("""
                INSERT INTO entradas
                (motorista, placa, telefone, fornecedor, destino, data_hora, porteiro)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            """, (
                var_motorista.get().strip().upper(),
                var_placa.get().strip().upper(),
                var_telefone.get().strip(),
                var_fornecedor.get().strip().upper(),
                var_destino.get().strip().upper(),
                datetime.now().strftime("%d/%m/%Y %H:%M"),
                var_porteiro.get().strip().upper()
            ))
            con.commit()
            entrada_id = cur.lastrowid

            registrar_log(
                acao="ENTRADA",
                entrada_id=entrada_id,
                placa=var_placa.get(),
                destino=var_destino.get(),
                porteiro=usuario_logado.get("nome") or var_porteiro.get(),
                detalhes=f"FORNECEDOR={var_fornecedor.get()} TEL={var_telefone.get()}"
            )

            set_status(f"Entrada registrada: {var_placa.get()} em {var_destino.get()}.", "success")
            limpar_campos()
            carregar_blocos()

        except Exception as e:
            messagebox.showerror("Erro ao registrar entrada", str(e))
            set_status("Erro ao registrar entrada.", "error")

    def registrar_saida(id_registro):
        if id_registro in baixas_em_andamento:
            return
        baixas_em_andamento.add(id_registro)

        try:
            cur.execute("""
                SELECT placa, motorista, destino, telefone, fornecedor, data_hora, porteiro
                FROM entradas
                WHERE id = ?
            """, (id_registro,))
            row = cur.fetchone()
            if not row:
                messagebox.showwarning("Não encontrado", "Registro não encontrado.")
                return

            placa_, motorista_, destino_, telefone_, fornecedor_, entrada_, porteiro_ = row

            if not confirmar_modal(
                "Confirmar baixa",
                "Deseja realmente registrar a SAÍDA deste veículo?\n\nEsta ação moverá o registro para o histórico.",
                texto_ok="REGISTRAR SAÍDA",
                texto_cancel="CANCELAR",
                tipo="warning",
                detalhes_linhas=[
                    f"PLACA: {placa_}",
                    f"MOTORISTA: {motorista_}",
                    f"DESTINO: {destino_}",
                    f"TEL: {telefone_}   |   FORN: {fornecedor_}",
                    f"ENTRADA: {entrada_}   |   PORTEIRO: {porteiro_}"
                ]
            ):
                return

            cur.execute("UPDATE entradas SET saida = ? WHERE id = ?",
                        (datetime.now().strftime("%d/%m/%Y %H:%M"), id_registro))
            con.commit()

            registrar_log(
                acao="BAIXA",
                entrada_id=id_registro,
                placa=placa_,
                destino=destino_,
                porteiro=usuario_logado.get("nome") or var_porteiro.get(),
                detalhes="REGISTRO DE SAÍDA"
            )

            set_status(f"Saída registrada: {placa_}.", "success")
            carregar_blocos()

        except Exception as e:
            messagebox.showerror("Erro ao registrar saída", str(e))
            set_status("Erro ao registrar saída.", "error")
        finally:
            baixas_em_andamento.discard(id_registro)

    def desfazer_baixa(id_registro):
        if usuario_logado.get("role") != "ADMIN":
            messagebox.showwarning("Acesso negado", "Apenas ADMIN pode desfazer baixa.")
            return

        if id_registro in desfazer_em_andamento:
            return
        desfazer_em_andamento.add(id_registro)

        try:
            cur.execute("""
                SELECT placa, motorista, destino, telefone, fornecedor, data_hora, saida, porteiro
                FROM entradas
                WHERE id = ?
            """, (id_registro,))
            row = cur.fetchone()
            if not row:
                messagebox.showwarning("Não encontrado", "Registro não encontrado.")
                return

            placa_, motorista_, destino_, telefone_, fornecedor_, entrada_, saida_, porteiro_ = row

            if not confirmar_modal(
                "Desfazer baixa",
                "Deseja DESFAZER a baixa deste veículo?\n\nO registro voltará para 'Entradas Ativas'.",
                texto_ok="DESFAZER",
                texto_cancel="CANCELAR",
                tipo="danger",
                detalhes_linhas=[
                    f"PLACA: {placa_}",
                    f"MOTORISTA: {motorista_}",
                    f"DESTINO: {destino_}",
                    f"TEL: {telefone_}   |   FORN: {fornecedor_}",
                    f"ENTRADA: {entrada_}   |   BAIXA: {saida_}",
                    f"PORTEIRO (REGISTRO): {porteiro_}",
                    f"ADMIN (DESFAZENDO): {usuario_logado.get('nome')}"
                ]
            ):
                return

            cur.execute("UPDATE entradas SET saida = NULL WHERE id = ?", (id_registro,))
            con.commit()

            registrar_log(
                acao="DESFAZER_BAIXA",
                entrada_id=id_registro,
                placa=placa_,
                destino=destino_,
                porteiro=usuario_logado.get("nome"),
                detalhes="BAIXA DESFEITA"
            )

            set_status(f"Baixa desfeita: {placa_}.", "success")
            carregar_blocos()

        except Exception as e:
            messagebox.showerror("Erro ao desfazer baixa", str(e))
            set_status("Erro ao desfazer baixa.", "error")
        finally:
            desfazer_em_andamento.discard(id_registro)

    # ================= HISTÓRICO / LOGS / RELATÓRIOS =================
    def abrir_historico():
        hist = tk.Toplevel(window)
        hist.title("Histórico de Baixas")
        hist.state("zoomed")
        hist.configure(bg=COLORS["bg"])

        container = ttk.Frame(hist, padding=14)
        container.pack(fill="both", expand=True)

        filtro = ttk.Frame(container, style="Card.TFrame", padding=12)
        filtro.pack(fill="x", pady=(0, 10))
        filtro.configure(relief="solid")
        filtro["borderwidth"] = 1

        ttk.Label(filtro, text="Filtros", style="CardTitle.TLabel").grid(row=0, column=0, sticky="w", pady=(0, 8))

        ttk.Label(filtro, text="Data da baixa (DD/MM/AAAA):", background=COLORS["card"]).grid(row=1, column=0, sticky="e")
        var_data = tk.StringVar()
        ttk.Entry(filtro, textvariable=var_data, width=16).grid(row=1, column=1, padx=6, sticky="w")

        ttk.Label(filtro, text="Empresa/Destino:", background=COLORS["card"]).grid(row=1, column=2, sticky="e")
        var_empresa = tk.StringVar()
        var_empresa.trace_add("write", lambda *a: forcar_maiusculo(var_empresa))
        ttk.Entry(filtro, textvariable=var_empresa, width=24).grid(row=1, column=3, padx=6, sticky="w")

        lista_card = ttk.Frame(container, style="Card.TFrame", padding=12)
        lista_card.pack(fill="both", expand=True)
        lista_card.configure(relief="solid")
        lista_card["borderwidth"] = 1

        ttk.Label(lista_card, text="Registros baixados", style="CardTitle.TLabel").pack(anchor="w", pady=(0, 10))

        canvas = tk.Canvas(lista_card, bg=COLORS["card"], highlightthickness=0)
        canvas.pack(side="left", fill="both", expand=True)
        sb = ttk.Scrollbar(lista_card, orient="vertical", command=canvas.yview)
        sb.pack(side="right", fill="y")
        canvas.configure(yscrollcommand=sb.set)

        frame_lista = ttk.Frame(canvas, style="Card.TFrame")
        win_id = canvas.create_window((0, 0), window=frame_lista, anchor="nw")

        def _scrollregion(event=None):
            canvas.configure(scrollregion=canvas.bbox("all"))

        frame_lista.bind("<Configure>", _scrollregion)

        def _resize(event):
            canvas.itemconfig(win_id, width=event.width)

        canvas.bind("<Configure>", _resize)

        def carregar_historico():
            for w in frame_lista.winfo_children():
                w.destroy()

            query = """
                SELECT id, motorista, placa, telefone, fornecedor, destino, data_hora, saida, porteiro
                FROM entradas
                WHERE saida IS NOT NULL
            """
            params = []

            if var_data.get().strip():
                query += " AND saida LIKE ?"
                params.append(f"%{var_data.get().strip()}%")

            if var_empresa.get().strip():
                query += " AND destino LIKE ?"
                params.append(f"%{var_empresa.get().strip().upper()}%")

            query += " ORDER BY saida DESC"

            cur.execute(query, params)
            rows = cur.fetchall()

            if not rows:
                ttk.Label(frame_lista, text="Nenhum registro encontrado.", background=COLORS["card"],
                          foreground=COLORS["muted"]).pack(anchor="w")
                return

            for r in rows:
                id_reg, motorista, placa, telefone, fornecedor, destino, entrada, saida, porteiro = r

                bloco = ttk.Frame(frame_lista, style="Card.TFrame", padding=10)
                bloco.pack(fill="x", pady=6)
                bloco.configure(relief="solid")
                bloco["borderwidth"] = 1

                ttk.Label(bloco, text=f"{placa}  •  {motorista}", background=COLORS["card"],
                          foreground=COLORS["primary"], font=("Segoe UI", 10, "bold")).pack(anchor="w")
                ttk.Label(bloco, text=f"TEL: {telefone}   |   FORN: {fornecedor}   |   DEST: {destino}",
                          background=COLORS["card"]).pack(anchor="w", pady=(6, 0))
                ttk.Label(bloco, text=f"ENTRADA: {entrada}   |   BAIXA: {saida}   |   PORTEIRO: {porteiro}",
                          background=COLORS["card"], foreground=COLORS["muted"]).pack(anchor="w", pady=(2, 0))

                tk.Button(
                    bloco, text="DESFAZER BAIXA", command=lambda i=id_reg: desfazer_baixa(i),
                    bd=0, padx=10, pady=6, bg=COLORS["warning"], fg="white",
                    font=("Segoe UI", 9, "bold"), cursor="hand2"
                ).pack(anchor="e", pady=(8, 0))

        tk.Button(
            filtro, text="FILTRAR", command=carregar_historico,
            bd=0, padx=10, pady=8, bg=COLORS["primary"], fg="white", font=("Segoe UI", 9, "bold")
        ).grid(row=1, column=4, padx=8)

        carregar_historico()

    def abrir_logs():
        logw = tk.Toplevel(window)
        logw.title("Logs do Sistema")
        logw.state("zoomed")
        logw.configure(bg=COLORS["bg"])

        container = ttk.Frame(logw, padding=14)
        container.pack(fill="both", expand=True)

        filtro = ttk.Frame(container, style="Card.TFrame", padding=12)
        filtro.pack(fill="x", pady=(0, 10))
        filtro.configure(relief="solid")
        filtro["borderwidth"] = 1

        ttk.Label(filtro, text="Filtros", style="CardTitle.TLabel").grid(row=0, column=0, sticky="w", pady=(0, 8))

        ttk.Label(filtro, text="Data (DD/MM/AAAA):", background=COLORS["card"]).grid(row=1, column=0, sticky="e")
        var_data = tk.StringVar()
        ttk.Entry(filtro, textvariable=var_data, width=16).grid(row=1, column=1, padx=6, sticky="w")

        ttk.Label(filtro, text="Placa:", background=COLORS["card"]).grid(row=1, column=2, sticky="e")
        var_placa_f = tk.StringVar()
        var_placa_f.trace_add("write", lambda *a: forcar_maiusculo(var_placa_f))
        ttk.Entry(filtro, textvariable=var_placa_f, width=12).grid(row=1, column=3, padx=6, sticky="w")

        ttk.Label(filtro, text="Porteiro:", background=COLORS["card"]).grid(row=1, column=4, sticky="e")
        var_porteiro_f = tk.StringVar()
        var_porteiro_f.trace_add("write", lambda *a: forcar_maiusculo(var_porteiro_f))
        ttk.Entry(filtro, textvariable=var_porteiro_f, width=18).grid(row=1, column=5, padx=6, sticky="w")

        ttk.Label(filtro, text="Ação:", background=COLORS["card"]).grid(row=1, column=6, sticky="e")
        var_acao = tk.StringVar(value="TODAS")
        combo_acao = ttk.Combobox(
            filtro, textvariable=var_acao, state="readonly",
            values=["TODAS", "ENTRADA", "BAIXA", "DESFAZER_BAIXA", "LOGOUT",
                    "USER_CREATE", "USER_RESET_PASS", "USER_ROLE", "USER_ACTIVE"],
            width=18
        )
        combo_acao.grid(row=1, column=7, padx=6, sticky="w")

        lista = ttk.Frame(container, style="Card.TFrame", padding=12)
        lista.pack(fill="both", expand=True)
        lista.configure(relief="solid")
        lista["borderwidth"] = 1

        ttk.Label(lista, text="Logs", style="CardTitle.TLabel").pack(anchor="w", pady=(0, 10))

        cols = ("id", "data_hora", "acao", "entrada_id", "placa", "destino", "porteiro", "detalhes")
        tree = ttk.Treeview(lista, columns=cols, show="headings")
        for c in cols:
            tree.heading(c, text=c.upper())

        tree.column("id", width=60, anchor="center")
        tree.column("data_hora", width=160, anchor="center")
        tree.column("acao", width=140, anchor="center")
        tree.column("entrada_id", width=90, anchor="center")
        tree.column("placa", width=90, anchor="center")
        tree.column("destino", width=140)
        tree.column("porteiro", width=140)
        tree.column("detalhes", width=500)

        tree.pack(fill="both", expand=True)

        sb_y = ttk.Scrollbar(lista, orient="vertical", command=tree.yview)
        sb_y.pack(side="right", fill="y")
        sb_x = ttk.Scrollbar(lista, orient="horizontal", command=tree.xview)
        sb_x.pack(side="bottom", fill="x")
        tree.configure(yscrollcommand=sb_y.set, xscrollcommand=sb_x.set)

        def carregar_logs():
            for i in tree.get_children():
                tree.delete(i)

            query = """
                SELECT id, data_hora, acao, entrada_id, placa, destino, porteiro, detalhes
                FROM logs
                WHERE 1=1
            """
            params = []

            if var_data.get().strip():
                query += " AND data_hora LIKE ?"
                params.append(f"%{var_data.get().strip()}%")

            if var_placa_f.get().strip():
                query += " AND placa LIKE ?"
                params.append(f"%{var_placa_f.get().strip().upper()}%")

            if var_porteiro_f.get().strip():
                query += " AND porteiro LIKE ?"
                params.append(f"%{var_porteiro_f.get().strip().upper()}%")

            if var_acao.get() != "TODAS":
                query += " AND acao = ?"
                params.append(var_acao.get())

            query += " ORDER BY id DESC"

            cur.execute(query, params)
            for r in cur.fetchall():
                tree.insert("", "end", values=r)

        tk.Button(
            filtro, text="FILTRAR", command=carregar_logs,
            bd=0, padx=10, pady=8, bg=COLORS["primary"], fg="white", font=("Segoe UI", 9, "bold")
        ).grid(row=1, column=8, padx=8)

        carregar_logs()

    # ===== Relatórios + Exportação CSV =====
    def _parse_dt(s: str):
        if not s:
            return None
        s = s.strip()
        for fmt in ("%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M", "%d/%m/%Y"):
            try:
                return datetime.strptime(s, fmt)
            except ValueError:
                pass
        return None

    def _parse_data_simples(s: str):
        if not s:
            return None
        s = s.strip()
        for fmt in ("%d/%m/%Y", "%d/%m/%Y %H:%M", "%d/%m/%Y %H:%M:%S"):
            try:
                return datetime.strptime(s, fmt)
            except ValueError:
                pass
        return None

    def abrir_relatorios():
        rel = tk.Toplevel(window)
        rel.title("Relatórios e Exportação")
        rel.state("zoomed")
        rel.configure(bg=COLORS["bg"])

        container = ttk.Frame(rel, padding=14)
        container.pack(fill="both", expand=True)

        filtro = ttk.Frame(container, style="Card.TFrame", padding=12)
        filtro.pack(fill="x", pady=(0, 10))
        filtro.configure(relief="solid")
        filtro["borderwidth"] = 1

        ttk.Label(filtro, text="Relatórios", style="CardTitle.TLabel").grid(row=0, column=0, sticky="w", pady=(0, 8))

        ttk.Label(filtro, text="Tipo de data:", background=COLORS["card"]).grid(row=1, column=0, sticky="e")
        var_tipo_data = tk.StringVar(value="ENTRADA")
        cb_tipo = ttk.Combobox(filtro, textvariable=var_tipo_data, state="readonly",
                               values=["ENTRADA", "BAIXA"], width=10)
        cb_tipo.grid(row=1, column=1, padx=6, sticky="w")

        ttk.Label(filtro, text="Início (DD/MM/AAAA):", background=COLORS["card"]).grid(row=1, column=2, sticky="e")
        var_ini = tk.StringVar()
        ttk.Entry(filtro, textvariable=var_ini, width=14).grid(row=1, column=3, padx=6, sticky="w")

        ttk.Label(filtro, text="Fim (DD/MM/AAAA):", background=COLORS["card"]).grid(row=1, column=4, sticky="e")
        var_fim = tk.StringVar()
        ttk.Entry(filtro, textvariable=var_fim, width=14).grid(row=1, column=5, padx=6, sticky="w")

        ttk.Label(filtro, text="Status:", background=COLORS["card"]).grid(row=1, column=6, sticky="e")
        var_status = tk.StringVar(value="TODOS")
        cb_status = ttk.Combobox(filtro, textvariable=var_status, state="readonly",
                                 values=["TODOS", "ATIVOS", "BAIXADOS"], width=10)
        cb_status.grid(row=1, column=7, padx=6, sticky="w")

        ttk.Label(filtro, text="Destino:", background=COLORS["card"]).grid(row=2, column=0, sticky="e")
        var_dest = tk.StringVar()
        var_dest.trace_add("write", lambda *a: forcar_maiusculo(var_dest))
        ttk.Entry(filtro, textvariable=var_dest, width=22).grid(row=2, column=1, padx=6, sticky="w")

        ttk.Label(filtro, text="Fornecedor:", background=COLORS["card"]).grid(row=2, column=2, sticky="e")
        var_forn = tk.StringVar()
        var_forn.trace_add("write", lambda *a: forcar_maiusculo(var_forn))
        ttk.Entry(filtro, textvariable=var_forn, width=22).grid(row=2, column=3, padx=6, sticky="w")

        ttk.Label(filtro, text="Porteiro:", background=COLORS["card"]).grid(row=2, column=4, sticky="e")
        var_port = tk.StringVar()
        var_port.trace_add("write", lambda *a: forcar_maiusculo(var_port))
        ttk.Entry(filtro, textvariable=var_port, width=18).grid(row=2, column=5, padx=6, sticky="w")

        resumo = ttk.Frame(container, style="Card.TFrame", padding=12)
        resumo.pack(fill="x", pady=(0, 10))
        resumo.configure(relief="solid")
        resumo["borderwidth"] = 1

        total_var = tk.StringVar(value="Total: 0")
        ativos_var = tk.StringVar(value="Ativos: 0")
        baixados_var = tk.StringVar(value="Baixados: 0")
        top_var = tk.StringVar(value="Por empresa: -")

        ttk.Label(resumo, textvariable=total_var, background=COLORS["card"], font=("Segoe UI", 10, "bold")).pack(side="left", padx=6)
        ttk.Label(resumo, textvariable=ativos_var, background=COLORS["card"]).pack(side="left", padx=16)
        ttk.Label(resumo, textvariable=baixados_var, background=COLORS["card"]).pack(side="left", padx=16)
        ttk.Label(resumo, textvariable=top_var, background=COLORS["card"], foreground=COLORS["muted"]).pack(side="left", padx=16)

        lista = ttk.Frame(container, style="Card.TFrame", padding=12)
        lista.pack(fill="both", expand=True)
        lista.configure(relief="solid")
        lista["borderwidth"] = 1

        cols = ("id", "motorista", "placa", "telefone", "fornecedor", "destino", "entrada", "saida", "porteiro")
        tree = ttk.Treeview(lista, columns=cols, show="headings")
        for c in cols:
            tree.heading(c, text=c.upper())

        tree.column("id", width=60, anchor="center")
        tree.column("motorista", width=200)
        tree.column("placa", width=90, anchor="center")
        tree.column("telefone", width=110, anchor="center")
        tree.column("fornecedor", width=180)
        tree.column("destino", width=140)
        tree.column("entrada", width=150, anchor="center")
        tree.column("saida", width=150, anchor="center")
        tree.column("porteiro", width=150)

        tree.pack(fill="both", expand=True)

        sb_y = ttk.Scrollbar(lista, orient="vertical", command=tree.yview)
        sb_y.pack(side="right", fill="y")
        sb_x = ttk.Scrollbar(lista, orient="horizontal", command=tree.xview)
        sb_x.pack(side="bottom", fill="x")
        tree.configure(yscrollcommand=sb_y.set, xscrollcommand=sb_x.set)

        dados_atual = []

        def carregar_relatorio():
            nonlocal dados_atual
            for i in tree.get_children():
                tree.delete(i)
            dados_atual = []

            ini = _parse_data_simples(var_ini.get())
            fim = _parse_data_simples(var_fim.get())
            if ini and fim and ini > fim:
                messagebox.showwarning("Período inválido", "Data inicial maior que data final.")
                return

            q = """
                SELECT id, motorista, placa, telefone, fornecedor, destino, data_hora, saida, porteiro
                FROM entradas
                WHERE 1=1
            """
            params = []

            st = var_status.get()
            if st == "ATIVOS":
                q += " AND saida IS NULL"
            elif st == "BAIXADOS":
                q += " AND saida IS NOT NULL"

            if var_dest.get().strip():
                q += " AND destino LIKE ?"
                params.append(f"%{var_dest.get().strip().upper()}%")

            if var_forn.get().strip():
                q += " AND fornecedor LIKE ?"
                params.append(f"%{var_forn.get().strip().upper()}%")

            if var_port.get().strip():
                q += " AND porteiro LIKE ?"
                params.append(f"%{var_port.get().strip().upper()}%")

            q += " ORDER BY id DESC"
            cur.execute(q, params)
            rows = cur.fetchall()

            tipo = var_tipo_data.get()
            filtrado = []
            for r in rows:
                id_, mot, pla, tel, forn, dest, entrada, saida, port = r
                dt_str = entrada if tipo == "ENTRADA" else saida
                dt = _parse_dt(dt_str) if dt_str else None

                if ini and dt and dt < ini:
                    continue
                if fim and dt:
                    fim_fechamento = fim.replace(hour=23, minute=59, second=59)
                    if dt > fim_fechamento:
                        continue

                if tipo == "BAIXA" and (var_ini.get().strip() or var_fim.get().strip()) and not saida:
                    continue

                filtrado.append(r)

            for r in filtrado:
                tree.insert("", "end", values=r)

            dados_atual = filtrado

            total = len(filtrado)
            ativos = sum(1 for r in filtrado if r[7] is None)
            baixados = total - ativos

            cont = {}
            for r in filtrado:
                d = (r[5] or "").strip().upper()
                cont[d] = cont.get(d, 0) + 1
            top = sorted(cont.items(), key=lambda x: x[1], reverse=True)[:3]
            top_txt = ", ".join([f"{k}:{v}" for k, v in top]) if top else "-"

            total_var.set(f"Total: {total}")
            ativos_var.set(f"Ativos: {ativos}")
            baixados_var.set(f"Baixados: {baixados}")
            top_var.set(f"Por empresa: {top_txt}")

        def exportar_xlsx():
            if not dados_atual:
                messagebox.showwarning("Nada para exportar", "Gere o relatório antes de exportar.")
                return

            arquivo = filedialog.asksaveasfilename(
                parent=rel,
                defaultextension=".xlsx",
                filetypes=[("Excel", "*.xlsx")],
                title="Salvar relatório como Excel"
            )
            if not arquivo:
                return

            try:
                wb = Workbook()
                ws = wb.active
                ws.title = "RELATORIO"

                headers = ["ID", "MOTORISTA", "PLACA", "TELEFONE", "FORNECEDOR", "DESTINO", "ENTRADA", "SAIDA",
                           "PORTEIRO"]

                # ===== Resumo (linhas no topo) =====
                titulo_fill = PatternFill("solid", fgColor="0B3D91")  # primary
                titulo_font = Font(color="FFFFFF", bold=True, size=12)
                subtitulo_font = Font(color="1F2937", bold=True)
                normal_font = Font(color="1F2937")
                left = Alignment(horizontal="left", vertical="center")
                center = Alignment(horizontal="center", vertical="center")

                # métricas básicas
                total = len(dados_atual)
                ativos = sum(1 for r in dados_atual if r[7] is None)  # saida
                baixados = total - ativos

                # filtros atuais (vêm dos campos do relatório)
                tipo_data = (var_tipo_data.get() if 'var_tipo_data' in locals() else "")
                ini = (var_ini.get().strip() if 'var_ini' in locals() else "")
                fim = (var_fim.get().strip() if 'var_fim' in locals() else "")
                status = (var_status.get() if 'var_status' in locals() else "")
                destino = (var_dest.get().strip() if 'var_dest' in locals() else "")
                fornecedor = (var_forn.get().strip() if 'var_forn' in locals() else "")
                porteiro = (var_port.get().strip() if 'var_port' in locals() else "")

                # Linha 1: Título
                ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
                c = ws.cell(row=1, column=1, value="RELATÓRIO - SISTEMA DE PORTARIA")
                c.fill = titulo_fill
                c.font = titulo_font
                c.alignment = center
                ws.row_dimensions[1].height = 22

                # Linha 2: Gerado em
                ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
                c = ws.cell(row=2, column=1, value=f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
                c.font = normal_font
                c.alignment = left

                # Linha 3: Período / Tipo data
                periodo_txt = f"Tipo de data: {tipo_data} | Período: {ini or '-'} até {fim or '-'} | Status: {status}"
                ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=len(headers))
                c = ws.cell(row=3, column=1, value=periodo_txt)
                c.font = subtitulo_font
                c.alignment = left

                # Linha 4: filtros textuais
                filtros_txt = f"Destino: {destino or '-'} | Fornecedor: {fornecedor or '-'} | Porteiro: {porteiro or '-'}"
                ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=len(headers))
                c = ws.cell(row=4, column=1, value=filtros_txt)
                c.font = subtitulo_font
                c.alignment = left

                # Linha 5: resumo numérico
                resumo_txt = f"Total: {total} | Ativos: {ativos} | Baixados: {baixados}"
                ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=len(headers))
                c = ws.cell(row=5, column=1, value=resumo_txt)
                c.font = subtitulo_font
                c.alignment = left

                # Linha 6: espaço
                ws.append([""] * len(headers))

                # ===== Cabeçalho da tabela =====
                header_fill = PatternFill("solid", fgColor="1E88E5")  # accent
                header_font = Font(color="FFFFFF", bold=True)
                header_alignment = Alignment(horizontal="center", vertical="center")

                ws.append(headers)
                header_row = ws.max_row

                for col in range(1, len(headers) + 1):
                    cell = ws.cell(row=header_row, column=col)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = header_alignment

                # ===== Dados =====
                for r in dados_atual:
                    ws.append(list(r))

                # Congela topo (até header da tabela)
                ws.freeze_panes = f"A{header_row + 1}"

                # Alinhamento
                for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, min_col=1, max_col=len(headers)):
                    for cell in row:
                        cell.alignment = Alignment(vertical="center")

                # Auto-ajuste simples das colunas
                for col in range(1, len(headers) + 1):
                    max_len = 0
                    col_letter = get_column_letter(col)
                    for cell in ws[col_letter]:
                        val = "" if cell.value is None else str(cell.value)
                        if len(val) > max_len:
                            max_len = len(val)
                    ws.column_dimensions[col_letter].width = min(max_len + 3, 55)

                wb.save(arquivo)
                messagebox.showinfo("Exportado", f"Relatório salvo em:\n{arquivo}")
                set_status("Relatório exportado em Excel.", "success")

            except Exception as e:
                messagebox.showerror("Erro ao exportar", str(e))
                set_status("Erro ao exportar Excel.", "error")

        btns = ttk.Frame(filtro, style="Card.TFrame")
        btns.grid(row=1, column=8, rowspan=2, padx=10, sticky="ns")

        tk.Button(btns, text="GERAR", command=carregar_relatorio, bd=0, padx=12, pady=8,
                  bg=COLORS["primary"], fg="white", font=("Segoe UI", 9, "bold")).pack(fill="x", pady=(0, 8))
        tk.Button(
            btns, text="EXPORTAR EXCEL", command=exportar_xlsx,
            bd=0, padx=12, pady=8,
            bg=COLORS["accent"], fg="white",
            font=("Segoe UI", 9, "bold")
        ).pack(fill="x")

        carregar_relatorio()

    # ===== Header Buttons =====
    btn_hist = header_button(header_btns, "Histórico", abrir_historico, COLORS["accent"])
    btn_logs = header_button(header_btns, "Logs", abrir_logs, COLORS["accent"])
    btn_rel = header_button(header_btns, "Relatórios", abrir_relatorios, COLORS["accent"])
    btn_users = header_button(header_btns, "Usuários", abrir_cadastro_usuarios, COLORS["accent"])
    btn_trocar = header_button(header_btns, "Trocar usuário", logout, COLORS["danger"])

    btn_hist.pack(side="left", padx=5)
    btn_logs.pack(side="left", padx=5)
    btn_rel.pack(side="left", padx=5)
    btn_users.pack(side="left", padx=5)
    btn_trocar.pack(side="left", padx=5)

    # ===== botões do lado esquerdo =====
    def registrar_entrada_btn():
        registrar_entrada()

    main_button(btns_card, "REGISTRAR ENTRADA", registrar_entrada_btn, COLORS["primary"]).pack(fill="x", pady=(0, 8))
    main_button(btns_card, "LIMPAR CAMPOS", limpar_campos, "#374151").pack(fill="x")

    # ================= BLOCOS (direita) =================
    blocos_card = ttk.Frame(right, style="Card.TFrame", padding=12)
    blocos_card.pack(fill="both", expand=True)
    blocos_card.configure(relief="solid")
    blocos_card["borderwidth"] = 1

    ttk.Label(blocos_card, text="Entradas Ativas", style="CardTitle.TLabel").pack(anchor="w", pady=(0, 10))
    ttk.Label(blocos_card, text="Organizado por empresa/destino.", background=COLORS["card"],
              foreground=COLORS["muted"]).pack(anchor="w", pady=(0, 10))

    blocos_container = ttk.Frame(blocos_card, style="Card.TFrame")
    blocos_container.pack(fill="both", expand=True)

    canvas_blocos = tk.Canvas(blocos_container, bg=COLORS["card"], highlightthickness=0)
    canvas_blocos.pack(side="left", fill="both", expand=True)

    scroll_y_blocos = ttk.Scrollbar(blocos_container, orient="vertical", command=canvas_blocos.yview)
    scroll_y_blocos.pack(side="right", fill="y")

    scroll_x_blocos = ttk.Scrollbar(blocos_card, orient="horizontal", command=canvas_blocos.xview)
    scroll_x_blocos.pack(fill="x", pady=(8, 0))

    canvas_blocos.configure(yscrollcommand=scroll_y_blocos.set, xscrollcommand=scroll_x_blocos.set)

    frame_blocos = ttk.Frame(canvas_blocos, style="Card.TFrame")
    canvas_window_id = canvas_blocos.create_window((0, 0), window=frame_blocos, anchor="nw")

    def _ajustar_scroll_blocos(event=None):
        canvas_blocos.configure(scrollregion=canvas_blocos.bbox("all"))

    frame_blocos.bind("<Configure>", _ajustar_scroll_blocos)

    def _on_canvas_configure(event):
        canvas_blocos.itemconfig(canvas_window_id)

    canvas_blocos.bind("<Configure>", _on_canvas_configure)

    def _cards_por_linha():
        w = window.winfo_width()
        return 3 if w >= 1700 else 2

    def carregar_blocos():
        for w in frame_blocos.winfo_children():
            w.destroy()

        filtro = var_busca.get().strip().upper()

        query = """
            SELECT id, motorista, placa, telefone, fornecedor, destino, data_hora, porteiro
            FROM entradas
            WHERE saida IS NULL
        """
        params = []
        if filtro:
            query += " AND (placa LIKE ? OR destino LIKE ?)"
            params.extend([f"%{filtro}%", f"%{filtro}%"])

        query += " ORDER BY destino, data_hora"
        cur.execute(query, params)
        registros = cur.fetchall()

        por_destino = {}
        for r in registros:
            por_destino.setdefault(r[5], []).append(r)

        destinos_ordenados = [d for d in EMPRESAS if d in por_destino]
        for d in por_destino.keys():
            if d not in destinos_ordenados:
                destinos_ordenados.append(d)

        CPL = _cards_por_linha()

        for col, destino in enumerate(destinos_ordenados):
            coluna = ttk.Frame(frame_blocos, style="Card.TFrame", padding=10)
            coluna.grid(row=0, column=col, padx=10, pady=10, sticky="n")
            coluna.configure(relief="solid")
            coluna["borderwidth"] = 1

            qtd = len(por_destino.get(destino, []))
            ttk.Label(coluna, text=f"{destino} ({qtd})", background=COLORS["card"],
                      foreground=COLORS["primary"], font=("Segoe UI", 10, "bold")).pack(anchor="w")

            ttk.Separator(coluna).pack(fill="x", pady=8)

            cards_area = ttk.Frame(coluna, style="Card.TFrame")
            cards_area.pack()

            itens = por_destino.get(destino, [])
            for idx, reg in enumerate(itens, start=1):
                id_reg, motorista, placa, telefone, fornecedor, _dest, data, porteiro = reg

                pos = idx - 1
                r = pos // CPL
                c = pos % CPL

                card = ttk.Frame(cards_area, style="Card.TFrame", padding=10)
                card.grid(row=r, column=c, padx=8, pady=8, sticky="n")
                card.configure(relief="solid")
                card["borderwidth"] = 1

                ttk.Label(card, text=f"{idx}) {placa}", background=COLORS["card"],
                          foreground=COLORS["text"], font=("Segoe UI", 10, "bold")).pack(anchor="w")
                ttk.Label(card, text=f"{motorista}", background=COLORS["card"],
                          foreground=COLORS["primary"]).pack(anchor="w")

                ttk.Label(card, text=f"TEL: {telefone}", background=COLORS["card"],
                          foreground=COLORS["muted"]).pack(anchor="w", pady=(6, 0))
                ttk.Label(card, text=f"FORN: {fornecedor}", background=COLORS["card"],
                          foreground=COLORS["muted"]).pack(anchor="w")
                ttk.Label(card, text=f"ENTRADA: {data}", background=COLORS["card"],
                          foreground=COLORS["muted"]).pack(anchor="w")
                ttk.Label(card, text=f"PORTEIRO: {porteiro}", background=COLORS["card"],
                          foreground=COLORS["muted"]).pack(anchor="w")

                tk.Button(
                    card, text="REGISTRAR SAÍDA", command=lambda i=id_reg: registrar_saida(i),
                    bd=0, padx=10, pady=8, bg=COLORS["success"], fg="white",
                    font=("Segoe UI", 9, "bold"), cursor="hand2"
                ).pack(fill="x", pady=(10, 0))

        _ajustar_scroll_blocos()

    # ===== Status bar =====
    status_label = tk.Label(root_container, textvariable=status_var, bg=COLORS["primary_dark"], fg="white",
                            padx=10, pady=8, anchor="w", font=("Segoe UI", 9))
    status_label.pack(fill="x")

    # ===== Inicialização =====
    abrir_login()
    carregar_blocos()
    set_status("Sistema pronto. Faça o cadastro de entradas.", "info")
    window.mainloop()


if __name__ == "__main__":
    iniciar_tela()
